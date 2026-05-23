import logging
import re

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def extract_party_a_contract(full_name: str) -> str:
    """从合同名称中提取甲方合同号。

    规则：
    1. 去掉 YW- 前缀（如有）
    2. 如果剩余部分以 SZ 开头 → 取到第二个 -（含）
    3. 否则 → 取到第一个 -（不含）
    """
    name = re.sub(r"^YW-", "", full_name)
    if name.startswith("SZ"):
        first_dash = name.index("-")
        second_dash = name.index("-", first_dash + 1)
        result = name[:second_dash]
    else:
        result = name.split("-")[0]
    logger.debug("合同名 '%s' → 甲方合同号 '%s'", full_name, result)
    return result


def load_contract_index(excel_path: str) -> dict[str, str]:
    """加载合同号索引表，返回 {乙方合同号: 甲方合同号} 映射。

    假设 Column A = 乙方合同号，Column H = 合同名称，从第2行开始读取。
    """
    logger.info("打开合同索引Excel: %s", excel_path)
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    index = {}
    skipped = 0
    for row in ws.iter_rows(min_row=2, max_col=8):
        party_b_id = row[0].value  # Column A
        contract_name = row[7].value  # Column H
        if party_b_id and contract_name:
            party_a_id = extract_party_a_contract(str(contract_name))
            index[str(party_b_id)] = party_a_id
        else:
            skipped += 1
    wb.close()
    logger.info("加载完成: %d条有效记录, %d条跳过(空行或缺少数据)", len(index), skipped)
    return index
