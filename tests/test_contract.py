import os
import tempfile

from openpyxl import Workbook

from contract import extract_party_a_contract, load_contract_index


def test_extract_non_sz():
    """非SZ开头的合同号，取到第一个-"""
    result = extract_party_a_contract("YW-7871Y00200172-0C1911-651400P")
    assert result == "7871Y00200172"


def test_extract_sz():
    """SZ开头的合同号，取到第二个-（含）"""
    result = extract_party_a_contract("YW-SZYG4502802-4-S292912F-3000P")
    assert result == "SZYG4502802-4"


def test_extract_no_prefix():
    """没有YW-前缀的情况，仍然正常处理"""
    result = extract_party_a_contract("7871Y00200172-0C1911-651400P")
    assert result == "7871Y00200172"


def test_load_contract_index():
    """从Excel加载合同号索引映射"""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "乙方合同号"
    ws["H1"] = "合同名称"
    ws["A2"] = "2100002601080A"
    ws["H2"] = "YW-7871Y00200172-0C1911-651400P"
    ws["A3"] = "2100002602020J"
    ws["H3"] = "YW-SZYG4502802-4-S292912F-3000P"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        tmp_path = f.name

    try:
        index = load_contract_index(tmp_path)
        assert index == {
            "2100002601080A": "7871Y00200172",
            "2100002602020J": "SZYG4502802-4",
        }
    finally:
        os.unlink(tmp_path)
