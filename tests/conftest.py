import os

import fitz
from openpyxl import Workbook


# 基于参考图中的真实数据
MOCK_CONTRACTS = [
    ("2100002601080A", "YW-7871Y00200172-0C1911-651400P"),
    ("2100002602020J", "YW-7871Y00200315L-092176-200P-0109"),
    ("2100002602040D", "YW-7871Y00200344H-1105-8040P-0106"),
    ("2100002602040F", "YW-7871Y00200343H-1105-246P-0106"),
    ("2100002602100D", "YW-SZYG4502802-4-S292912F-3000P"),
]

MOCK_INVOICES = [
    {
        "party_a_id": "7871Y00200172",
        "party_b_id": "2100002601080A",
        "billing_qty": 1265,
        "invoice_no": "",  # 空，由OCR填充
        "invoice_date": "",  # 空，由OCR填充
        "tax_amount": 256308.48,
        "total_amount": 2227912.18,
    },
    {
        "party_a_id": "7871Y00200315L",
        "party_b_id": "2100002602020J",
        "billing_qty": 925,
        "invoice_no": "26317000001473818224",
        "invoice_date": "2026/05/19",
        "tax_amount": 187419.25,
        "total_amount": 1629105.75,
    },
    {
        "party_a_id": "7871Y00200344H",
        "party_b_id": "2100002602040D",
        "billing_qty": 1127,
        "invoice_no": "",  # 空，由OCR填充
        "invoice_date": "",  # 空，由OCR填充
        "tax_amount": 228347.56,
        "total_amount": 1984867.22,
    },
]

# PDF中实际包含的发票号（OCR会识别这些，填入Excel中为空的行）
PDF_INVOICE_NOS = [
    "26317000001473818247",
    "26317000001473818224",
    "26317000001473818221",
]

PDF_INVOICE_DATES = [
    "2026/05/19",
    "2026/05/19",
    "2026/05/19",
]


def _create_contract_excel(path: str):
    """创建合同号索引Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "合同汇总"
    ws["A1"] = "乙方合同号"
    ws["H1"] = "合同名称"
    ws["A2"] = "summary information"

    for i, (party_b, full_name) in enumerate(MOCK_CONTRACTS, start=3):
        ws[f"A{i}"] = party_b
        ws[f"H{i}"] = full_name

    wb.save(path)
    wb.close()


def _create_invoice_excel(path: str):
    """创建发票验证Excel"""
    wb = Workbook()
    ws = wb.active
    # 表头 (Row 1-3)
    ws.cell(row=1, column=32, value="Fields in this area ca")
    ws.cell(row=2, column=32, value="PO No.")
    ws.cell(row=2, column=70, value="Billing Qty*")
    ws.cell(row=2, column=79, value="Invoice No*")
    ws.cell(row=2, column=80, value="Invoice Date*")
    ws.cell(row=2, column=81, value="Tax Amount*")
    ws.cell(row=2, column=82, value="Invoice Amount (Incl. Tax)*")
    ws.cell(row=3, column=32, value="PO号")
    ws.cell(row=3, column=70, value="本次开票数量")
    ws.cell(row=3, column=79, value="发票号")
    ws.cell(row=3, column=80, value="发票日期")
    ws.cell(row=3, column=81, value="税金金额")
    ws.cell(row=3, column=82, value="含税金额")

    # 数据行 (Row 4+)
    for i, inv in enumerate(MOCK_INVOICES, start=4):
        ws.cell(row=i, column=32, value=inv["party_a_id"])  # AF
        ws.cell(row=i, column=70, value=inv["billing_qty"])  # BZ
        ws.cell(row=i, column=79, value=inv["invoice_no"])   # CA
        ws.cell(row=i, column=80, value=inv["invoice_date"]) # CB
        ws.cell(row=i, column=81, value=inv["tax_amount"])   # CC
        ws.cell(row=i, column=82, value=inv["total_amount"]) # CD

    wb.save(path)
    wb.close()


def _create_mock_pdf(path: str, party_b_id: str, invoice_no: str,
                     invoice_date: str, tax_amount: str, total_amount: str):
    """创建包含发票内容的mock PDF（有文本层）"""
    doc = fitz.open()
    page = doc.new_page()

    # 将日期格式 2026/05/19 转为 2026年05月19日
    date_parts = invoice_date.split("/")
    date_cn = f"{date_parts[0]}年{date_parts[1]}月{date_parts[2]}日" if len(date_parts) == 3 else invoice_date

    # 构造发票文本（中文，匹配OCR正则）
    text = (
        f"发票号码：{invoice_no}\n"
        f"开票日期：{date_cn}\n"
        f"数量：10000\n"
        f"金额：{total_amount}\n"
        f"税率：13%\n"
        f"税额：{tax_amount}\n"
        f"价税合计：{total_amount}\n"
        f"备注：\n"
        f"{party_b_id}\n"
    )

    # 尝试使用系统中文字体
    font_candidates = [
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
    ]
    font_file = None
    for fp in font_candidates:
        if os.path.exists(fp):
            font_file = fp
            break

    if font_file:
        page.insert_text((50, 50), text, fontsize=10, fontname="china-ss", fontfile=font_file)
    else:
        # fallback: 使用默认字体（中文可能显示为方块，但文本层仍有内容）
        page.insert_text((50, 50), text, fontsize=10)

    doc.save(path)
    doc.close()


def generate_all_test_data(output_dir: str):
    """生成所有测试数据"""
    os.makedirs(output_dir, exist_ok=True)
    pdf_dir = os.path.join(output_dir, "pdfs")
    os.makedirs(pdf_dir, exist_ok=True)

    # 生成合同号索引Excel
    contract_excel = os.path.join(output_dir, "合同号索引表.xlsx")
    _create_contract_excel(contract_excel)

    # 生成发票验证Excel
    invoice_excel = os.path.join(output_dir, "发票验证表.xlsx")
    _create_invoice_excel(invoice_excel)

    # 生成Type 1 PDF
    for i, inv in enumerate(MOCK_INVOICES):
        filename = f"{inv['party_b_id']}_引望智能技术有限公司_20260519.pdf"
        _create_mock_pdf(
            os.path.join(pdf_dir, filename),
            party_b_id=inv["party_b_id"],
            invoice_no=PDF_INVOICE_NOS[i],
            invoice_date=PDF_INVOICE_DATES[i],
            tax_amount=str(inv["tax_amount"]),
            total_amount=str(inv["total_amount"]),
        )

    # 生成Type 2 PDF
    for i, inv in enumerate(MOCK_INVOICES):
        filename = f"dzfp_{PDF_INVOICE_NOS[i]}_引望智能技术有限公司_20260519.pdf"
        _create_mock_pdf(
            os.path.join(pdf_dir, filename),
            party_b_id=inv["party_b_id"],
            invoice_no=PDF_INVOICE_NOS[i],
            invoice_date=PDF_INVOICE_DATES[i],
            tax_amount=str(inv["tax_amount"]),
            total_amount=str(inv["total_amount"]),
        )

    return {
        "contract_excel": contract_excel,
        "invoice_excel": invoice_excel,
        "pdf_dir": pdf_dir,
    }
