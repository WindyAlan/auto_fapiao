from openpyxl import load_workbook


# Excel列号映射（基于列字母）
COLUMN_MAP = {
    "party_a_id": "AF",    # 甲方合同号
    "billing_qty": "BZ",   # 本次开票数量
    "invoice_no": "CA",    # 发票号
    "invoice_date": "CB",  # 发票日期
    "tax_amount": "CC",    # 税金金额
    "total_amount": "CD",  # 含税金额
}

# 数据起始行
DATA_START_ROW = 4


def load_workbook_rw(path: str):
    """加载Excel工作簿（读写模式）"""
    return load_workbook(path)


def get_column_index(col_letter: str) -> int:
    """将列字母转换为1-based列号，如 'AF' → 32"""
    result = 0
    for c in col_letter:
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def read_invoice_rows(ws) -> list[dict]:
    """从工作表中读取发票数据行（从DATA_START_ROW开始）"""
    max_col = get_column_index("CD")  # 确保读到CD列
    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=max_col):
        party_a_id = row[get_column_index("AF") - 1].value
        if not party_a_id:
            continue
        rows.append({
            "party_a_id": str(party_a_id),
            "billing_qty": str(row[get_column_index("BZ") - 1].value or ""),
            "invoice_no": str(row[get_column_index("CA") - 1].value or ""),
            "invoice_date": str(row[get_column_index("CB") - 1].value or ""),
            "tax_amount": str(row[get_column_index("CC") - 1].value or ""),
            "total_amount": str(row[get_column_index("CD") - 1].value or ""),
        })
    return rows


def write_cell(ws, row_idx: int, col_letter: str, value):
    """写入单元格"""
    col_idx = get_column_index(col_letter)
    ws.cell(row=row_idx, column=col_idx, value=value)
