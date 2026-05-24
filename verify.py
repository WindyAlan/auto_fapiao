import logging
import os
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

from excel_utils import COLUMN_MAP, get_column_index, read_invoice_rows
from ocr import extract_invoice_fields, extract_text_from_pdf, get_ocr_confidence

logger = logging.getLogger(__name__)

CONFIDENCE_THRESHOLD = 0.8


@dataclass
class FieldDiff:
    field_name: str
    excel_value: str
    ocr_value: str
    confidence: float
    fixed: bool


@dataclass
class VerifyResult:
    pdf_file: str
    party_a_id: str
    diffs: list[FieldDiff] = field(default_factory=list)
    filled: list[str] = field(default_factory=list)  # OCR填充的字段描述
    needs_manual: bool = False


def compare_fields(excel_row: dict, ocr_fields: dict, confidence: float) -> list[FieldDiff]:
    """比对Excel行与OCR识别结果，返回差异列表"""
    field_map = {
        "invoice_no": "发票号",
        "invoice_date": "发票日期",
        "tax_amount": "税金金额",
        "total_amount": "含税金额",
    }
    diffs = []
    for key, label in field_map.items():
        excel_val = str(excel_row.get(key, "")).strip()
        ocr_val = str(ocr_fields.get(key, "")).strip()
        if excel_val and ocr_val and excel_val != ocr_val:
            logger.debug("字段 '%s' 不一致: Excel='%s', OCR='%s'", label, excel_val, ocr_val)
            diffs.append(FieldDiff(
                field_name=label,
                excel_value=excel_val,
                ocr_value=ocr_val,
                confidence=confidence,
                fixed=False,
            ))
    return diffs


def resolve_party_a_id_from_filename(filename: str) -> str | None:
    """从已重命名的文件名中提取甲方合同号"""
    m = re.match(r"^(.+?)-(\d+[A-Z]+)_", filename)
    if m:
        return m.group(1)
    return None


def verify_invoices(pdf_dir: str, excel_path: str) -> tuple[list[VerifyResult], str]:
    """校验目录中所有PDF发票与Excel数据，并将OCR结果填入新的Excel文件。

    Returns:
        (结果列表, 输出Excel路径)
    """
    # 生成输出文件名
    base, ext = os.path.splitext(excel_path)
    output_excel = f"{base}_Verified{ext}"

    logger.info("打开验证Excel: %s", excel_path)
    wb = load_workbook(excel_path)
    ws = wb.active
    excel_rows = read_invoice_rows(ws)
    logger.info("Excel中读取到 %d 行发票数据", len(excel_rows))

    # 按甲方合同号索引Excel行（同时记录行号用于回写）
    excel_by_party_a = {r["party_a_id"]: r for r in excel_rows}

    pdf_files = [f for f in sorted(os.listdir(pdf_dir)) if f.lower().endswith(".pdf")]
    logger.info("发现 %d 个PDF文件", len(pdf_files))

    results = []
    filled_count = 0
    for filename in pdf_files:
        party_a_id = resolve_party_a_id_from_filename(filename)
        if not party_a_id:
            logger.warning("文件名格式无法识别甲方合同号: %s", filename)
            results.append(VerifyResult(
                pdf_file=filename, party_a_id="",
                needs_manual=True,
            ))
            continue

        excel_row = excel_by_party_a.get(party_a_id)
        if not excel_row:
            logger.warning("Excel中未找到甲方合同号: %s (文件: %s)", party_a_id, filename)
            results.append(VerifyResult(
                pdf_file=filename, party_a_id=party_a_id,
                needs_manual=True,
            ))
            continue

        pdf_path = os.path.join(pdf_dir, filename)
        text = extract_text_from_pdf(pdf_path)
        ocr_fields = extract_invoice_fields(text)
        confidence = get_ocr_confidence(pdf_path)

        diffs = compare_fields(excel_row, ocr_fields, confidence)

        # 将OCR识别的发票号和发票日期填入Excel（如果原表为空）
        filled = []
        row_idx = excel_row.get("_row_idx")
        if row_idx:
            for field_key, col_letter, label in [
                ("invoice_no", "CA", "发票号"),
                ("invoice_date", "CB", "发票日期"),
            ]:
                ocr_val = ocr_fields.get(field_key, "")
                excel_val = str(excel_row.get(field_key, "")).strip()
                if ocr_val and not excel_val:
                    col_idx = get_column_index(col_letter)
                    ws.cell(row=row_idx, column=col_idx, value=ocr_val)
                    filled_count += 1
                    filled.append(f"{label}={ocr_val}")
                    logger.info("填充: %s %s = %s", party_a_id, field_key, ocr_val)

        # 标记需要手动核查的差异
        needs_manual = bool(diffs)
        for diff in diffs:
            if diff.confidence <= CONFIDENCE_THRESHOLD:
                logger.warning("需手动核查: %s %s, Excel='%s' OCR='%s' (置信度=%.3f)",
                               party_a_id, diff.field_name, diff.excel_value, diff.ocr_value,
                               confidence)

        if not diffs:
            logger.debug("✓ %s: 所有字段正确", party_a_id)

        results.append(VerifyResult(
            pdf_file=filename,
            party_a_id=party_a_id,
            diffs=diffs,
            filled=filled,
            needs_manual=needs_manual,
        ))

    # 保存为新Excel文件
    # 如果原文件是 .xlsm，同时保存 .xlsm 和 .xlsx 两份
    base, ext = os.path.splitext(excel_path)
    if ext.lower() == ".xlsm":
        xlsm_path = f"{base}_Verified.xlsm"
        xlsx_path = f"{base}_Verified.xlsx"
        wb.save(xlsm_path)
        wb.close()
        # 重新加载 .xlsm 再另存为干净的 .xlsx
        wb2 = load_workbook(xlsm_path)
        wb2.save(xlsx_path)
        wb2.close()
        logger.info("验证完成，填充了 %d 个字段", filled_count)
        logger.info("  .xlsm 版本: %s", xlsm_path)
        logger.info("  .xlsx 版本: %s", xlsx_path)
        return results, xlsx_path
    else:
        wb.save(output_excel)
        wb.close()
        logger.info("验证完成，填充了 %d 个字段，已保存: %s", filled_count, output_excel)
        return results, output_excel
